import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configuración visual ultra simplificada para el usuario final
st.set_page_config(page_title="UMA - Convertidor de Notas", page_icon="📝", layout="centered")

st.markdown("<h2 style='text-align: center; color: #1E3A8A; font-family: Arial; font-weight: bold;'>🏢 UNIVERSIDAD MODULAR ABIERTA</h2>", unsafe_allow_html=True)
st.markdown("<h3 style='text-align: center; color: #4B5563; font-family: Arial;'>📊 Convertidor de Listados UONLINE</h3>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center; color: #7C3AED; font-weight: bold;'>Uso exclusivo para Control de Asistencia y Cuadro de Notas</p>", unsafe_allow_html=True)
st.markdown("---")

# Única opción en pantalla
pdf_adjunto = st.file_uploader("📥 ARRASTRE AQUÍ EL PDF DE LA UNIVERSIDAD O HAGA CLIC PARA BUSCARLO", type=["pdf"])

if pdf_adjunto is not None:
    with st.spinner("⏳ Leyendo el listado de alumnos... Por favor espere."):
        try:
            texto_completo = ""
            alumnos_detectados = []
            
            # 1. Leer el PDF con pdfplumber extrayendo las celdas exactas del reporte de la UMA
            with pdfplumber.open(pdf_adjunto) as pdf:
                for page in pdf.pages:
                    texto_completo += page.extract_text() + "\n"
                    tablas = page.extract_tables()
                    for tabla in tablas:
                        for fila in tabla:
                            # Limpiar las celdas de valores nulos o vacíos
                            fila_limpia = [str(c).strip() for c in fila if c is not None]
                            if not fila_limpia:
                                continue
                            
                            # Buscar el patrón de Carnet de la UMA (Ej: CB252100307 o SA252100175)
                            carnet = None
                            nombre_completo = None
                            
                            for celda in fila_limpia:
                                match_carnet = re.search(r'([A-Z]{2}\d{9})', celda)
                                if match_carnet:
                                    carnet = match_carnet.group(1)
                                    break
                            
                            # Si encontramos un carnet en esta fila, buscamos el nombre del alumno
                            if carnet:
                                for celda in fila_limpia:
                                    # El nombre está en la celda que tiene texto largo en mayúsculas y no es el encabezado
                                    celda_clean = celda.replace("\n", " ").strip()
                                    if len(celda_clean) > 10 and carnet not in celda_clean and "NOMBRE" not in celda_clean and "Firma" not in celda_clean:
                                        # Limpiar números que a veces se pegan al inicio o final debido al formato del reporte
                                        nombre_completo = re.sub(r'^\d+\s*|\s*\d+$', '', celda_clean).strip()
                                        break
                                
                                if carnet and nombre_completo:
                                    alumnos_detectados.append({
                                        "carnet_raw": carnet,
                                        "nombre_completo": nombre_completo
                                    })

            # 2. Extracción de Datos de la Materia desde la Cabecera Real del PDF
            facultad_m = re.search(r'Facultad\n\s*([^\n]+)', texto_completo)
            asignatura_m = re.search(r'Asignatura\n\s*([^\n]+)', texto_completo)
            ciclo_m = re.search(r'Desde:\s*\",?\"(CICLO\s*\d+-\d+)', texto_completo) or re.search(r'(CICLO\s*\d+-\d+)', texto_completo)
            aula_m = re.search(r'\"Aula\n\",\"([^\"]+)\"', texto_completo) or re.search(r'Aula\n\s*([^\n]+)', texto_completo)
            horario_m = re.search(r'\"Horario\n\",\"([^\"]+)\"', texto_completo) or re.search(r'Horario\n\s*([^\n]+)', texto_completo)
            dias_m = re.search(r'\"Días\n\",\"([^\"]+)\"', texto_completo) or re.search(r'Días\n\s*([^\n]+)', texto_completo)
            docente_m = re.search(r'Docente\n\s*([^\n]+)', texto_completo)

            f_txt = facultad_m.group(1).strip().upper() if facultad_m else "FACULTAD DE CIENCIAS ECONÓMICAS"
            a_txt = asignatura_m.group(1).strip().upper() if asignatura_m else "SISTEMAS OPERATIVOS"
            c_txt = "01-2026"  # Valor por defecto limpio según tu plantilla
            if ciclo_m:
                c_clean = ciclo_m.group(1).replace("CICLO", "").strip()
                if c_clean: c_txt = c_clean
                
            d_txt = docente_m.group(1).strip().upper() if docente_m else "LIC. OMAR ALBERTO LINARES DEL CID"
            
            # Formatear el horario exacto combinando Día + Horas limpias
            dia = dias_m.group(1).replace("\n", "").strip().upper() if dias_m else "JUEVES"
            horas = horario_m.group(1).replace("\n", "").strip().upper() if horario_m else "13:00 P.M. - 16:40 P.M."
            horas_clean = horas.replace(" P.M.", "").replace("P.M.", "")
            h_txt = f"{dia} DE {horas_clean} P.M."

            # 3. Procesar y Separar Apellidos y Nombres (Regla de los 2 primeros elementos)
            lista_final = []
            vistos = set()
            
            for al in alumnos_detectados:
                correo = f"{al['carnet_raw'].lower()}@uma.edu.sv"
                if correo in vistos:
                    continue
                vistos.add(correo)
                
                partes = [p for p in al['nombre_completo'].split(" ") if p]
                if len(partes) >= 3:
                    apellidos = f"{partes[0]} {partes[1]}".upper()
                    nombres = " ".join(partes[2:]).upper()
                else:
                    apellidos = al['nombre_completo'].upper()
                    nombres = ""
                
                lista_final.append({
                    "correo": correo,
                    "apellidos": apellidos,
                    "nombres": nombres
                })

            # Ordenar por Apellido Alfabéticamente tal como lo requiere el cuadro final
            lista_final = sorted(lista_final, key=lambda x: x['apellidos'])

            if not lista_final:
                st.error("❌ El formato de este PDF no coincide con el Control de Asistencia de la UMA. Por favor verifica el archivo.")
            else:
                # 4. Crear el Excel con el Formato Institucional Exacto y Fórmulas Activas
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Notas"

                # Estilos visuales idénticos a tu plantilla original
                f_tit = Font(name='Arial', size=10, bold=True)
                f_norm = Font(name='Arial', size=10)
                f_enc = Font(name='Arial', size=9, bold=True)
                fill_gray = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                thin_b = Border(left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                                top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'))

                # Escribir Cabecera en las filas de la 1 a la 8
                ws['A1'] = "UNIVERSIDAD MODULAR ABIERTA"
                ws['A2'] = f"FACULTAD : {f_txt}"
                ws['A3'] = "CENTRO : SEDE SONSONATE"
                ws['A4'] = "CUADRO DE EVALUACION"
                ws['A5'] = f"ASIGNATURA : {a_txt}"
                ws['A6'] = f"CICLO : {c_txt}"
                ws['A7'] = f"HORARIO : {h_txt}"
                ws['A8'] = f"DOCENTE : {d_txt}"
                ws['E8'], ws['K8'], ws['Q8'] = "PERIODO 1", "PERIODO 2", "PERIODO 3"

                for r in range(1, 9):
                    ws.cell(row=r, column=1).font = f_tit

                # Encabezados de Columnas Oficiales (Fila 10)
                headers = ["N", "CARNET", "APELLIDOS", "NOMBRES", "LAB1", "0.4", "PAR1", "0.6", "PARC", "P.N1.", 
                           "LAB2", "0.4", "PAR2", "0.6", "PARC", "P.N2.", "LAB3", "0.4", "PAR3", "0.6", "PARC", "P.N3.", "PROM. FINAL", "OBSERVACIONES"]
                for c_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=10, column=c_idx, value=h)
                    cell.font = f_enc; cell.fill = fill_gray; cell.border = thin_b
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Multiplicadores de periodo establecidos en la Fila 11
                ws.cell(row=11, column=10, value=0.3)
                ws.cell(row=11, column=16, value=0.3)
                ws.cell(row=11, column=22, value=0.4)
                ws.cell(row=11, column=23, value="FINAL")
                for c in range(1, 25):
                    ws.cell(row=11, column=c).font = f_enc; ws.cell(row=11, column=c).border = thin_b

                # Volcar a los alumnos y programar las celdas con fórmulas matemáticas automatizadas
                for idx, al in enumerate(lista_final, 1):
                    r = 11 + idx
                    ws.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
                    ws.cell(row=r, column=2, value=al['correo'])
                    ws.cell(row=r, column=3, value=al['apellidos'])
                    ws.cell(row=r, column=4, value=al['nombres'])
                    
                    # Ponderaciones fijas por fila de notas individuales
                    ws.cell(row=r, column=6, value=0.4)
                    ws.cell(row=r, column=8, value=0.6)
                    ws.cell(row=r, column=12, value=0.4)
                    ws.cell(row=r, column=14, value=0.6)
                    ws.cell(row=r, column=18, value=0.4)
                    ws.cell(row=r, column=20, value=0.6)
                    
                    # Fórmulas de Excel vivas
                    ws.cell(row=r, column=9, value=f"=E{r}*F{r}+G{r}*H{r}")   # PARC 1
                    ws.cell(row=r, column=10, value=f"=I{r}*J11")            # P.N1 (30%)
                    ws.cell(row=r, column=15, value=f"=K{r}*L{r}+M{r}*N{r}") # PARC 2
                    ws.cell(row=r, column=16, value=f"=O{r}*P11")            # P.N2 (30%)
                    ws.cell(row=r, column=21, value=f"=Q{r}*R{r}+S{r}*T{r}") # PARC 3
                    ws.cell(row=r, column=22, value=f"=U{r}*V11")            # P.N3 (40%)
                    ws.cell(row=r, column=23, value=f"=J{r}+P{r}+V{r}")      # PROM. FINAL TOTAL
                    
                    # Formatear fuentes y bordes de la fila
                    for c in range(1, 25):
                        ws.cell(row=r, column=c).font = f_norm; ws.cell(row=r, column=c).border = thin_b

                # Autoajustar el ancho de las columnas
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 9)

                wb.save(output)
                excel_data = output.getvalue()
                
                # --- RESPUESTA ÚNICA PARA EL USUARIO ---
                st.markdown("<br>", unsafe_allow_html=True)
                st.success(f"🎉 ¡Hecho! Se encontraron {len(lista_final)} alumnos ordenados alfabéticamente.")
                
                st.download_button(
                    label="🟢 DESCARGAR ARCHIVO EXCEL DE NOTAS",
                    data=excel_data,
                    file_name=f"Notas_{a_txt.replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        except Exception as e:
            st.error(f"Error técnico al leer el PDF. Asegúrate de que no esté corrupto.")